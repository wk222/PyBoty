"""Data source connectors for workflow nodes.

Provides pluggable data fetching from external sources:
  - HTTP/REST API (GET/POST with auth)
  - Database query (SQLite / PostgreSQL)
  - File system (CSV / JSON / Excel)
  - RSS / Atom feeds

Each connector returns structured data (list of dicts) that downstream
workflow nodes (Transform, LLM, Notify) can consume.
"""

from __future__ import annotations

import csv
import json
import logging
import os
import time
from dataclasses import dataclass, field
from typing import Any

logger = logging.getLogger(__name__)


@dataclass
class DataSourceConfig:
    """Configuration for a data source connector."""

    source_type: str  # "http", "database", "file", "rss"
    url: str = ""
    method: str = "GET"
    headers: dict[str, str] = field(default_factory=dict)
    body: dict[str, Any] | None = None
    auth_type: str = ""  # "", "bearer", "basic", "api_key"
    auth_value: str = ""
    query: str = ""  # SQL or JMESPath
    file_path: str = ""
    timeout: float = 30.0
    retry_count: int = 2
    extract_path: str = ""  # JSONPath-like extraction (e.g. "data.items")

    def to_dict(self) -> dict[str, Any]:
        return {k: v for k, v in self.__dict__.items() if v}

    @classmethod
    def from_dict(cls, data: dict[str, Any]) -> "DataSourceConfig":
        return cls(**{k: v for k, v in data.items() if k in cls.__dataclass_fields__})


@dataclass
class FetchResult:
    """Result of a data fetch operation."""

    success: bool
    data: list[dict[str, Any]] = field(default_factory=list)
    raw: str = ""
    record_count: int = 0
    elapsed_ms: float = 0
    error: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)


def _extract_nested(data: Any, path: str) -> Any:
    """Extract nested value from data using dot notation."""
    if not path:
        return data
    parts = path.split(".")
    current = data
    for part in parts:
        if isinstance(current, dict):
            current = current.get(part)
        elif isinstance(current, list) and part.isdigit():
            idx = int(part)
            current = current[idx] if idx < len(current) else None
        else:
            return None
    return current


def _normalize_to_records(data: Any) -> list[dict[str, Any]]:
    """Normalize various data shapes into a list of dicts."""
    if isinstance(data, list):
        return [item if isinstance(item, dict) else {"value": item} for item in data]
    if isinstance(data, dict):
        return [data]
    return [{"value": data}]


def fetch_http(config: DataSourceConfig) -> FetchResult:
    """Fetch data from an HTTP API."""
    import urllib.request
    import urllib.error

    start = time.time()
    try:
        headers = dict(config.headers)
        if config.auth_type == "bearer" and config.auth_value:
            headers["Authorization"] = f"Bearer {config.auth_value}"
        elif config.auth_type == "api_key" and config.auth_value:
            headers["X-API-Key"] = config.auth_value

        body_bytes = None
        if config.body and config.method.upper() in ("POST", "PUT", "PATCH"):
            body_bytes = json.dumps(config.body).encode("utf-8")
            headers.setdefault("Content-Type", "application/json")

        req = urllib.request.Request(
            config.url,
            data=body_bytes,
            headers=headers,
            method=config.method.upper(),
        )

        with urllib.request.urlopen(req, timeout=config.timeout) as resp:
            raw = resp.read().decode("utf-8")

        data = json.loads(raw)
        if config.extract_path:
            data = _extract_nested(data, config.extract_path)

        records = _normalize_to_records(data)
        elapsed = (time.time() - start) * 1000

        return FetchResult(
            success=True,
            data=records,
            raw=raw[:5000],
            record_count=len(records),
            elapsed_ms=elapsed,
            metadata={"url": config.url, "status": 200},
        )
    except Exception as exc:
        return FetchResult(
            success=False,
            error=str(exc),
            elapsed_ms=(time.time() - start) * 1000,
        )


def fetch_file(config: DataSourceConfig) -> FetchResult:
    """Fetch data from a local file (CSV, JSON, Excel)."""
    start = time.time()
    path = config.file_path
    if not os.path.isfile(path):
        return FetchResult(success=False, error=f"File not found: {path}")

    try:
        ext = os.path.splitext(path)[1].lower()

        if ext == ".json":
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            records = _normalize_to_records(data)

        elif ext == ".csv":
            with open(path, encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                records = list(reader)

        elif ext in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    headers_row = [str(c) if c else f"col_{i}" for i, c in enumerate(rows[0])]
                    records = [
                        dict(zip(headers_row, [v for v in row], strict=False))
                        for row in rows[1:]
                    ]
                else:
                    records = []
                wb.close()
            except ImportError:
                return FetchResult(success=False, error="Excel requires openpyxl: pip install openpyxl")

        else:
            with open(path, encoding="utf-8") as f:
                text = f.read()
            records = [{"content": text}]

        elapsed = (time.time() - start) * 1000
        return FetchResult(
            success=True,
            data=records,
            record_count=len(records),
            elapsed_ms=elapsed,
            metadata={"file": path, "format": ext},
        )
    except Exception as exc:
        return FetchResult(
            success=False,
            error=str(exc),
            elapsed_ms=(time.time() - start) * 1000,
        )


def fetch_database(config: DataSourceConfig) -> FetchResult:
    """Execute a SQL query and return results."""
    import sqlite3

    start = time.time()
    try:
        db_path = config.url or config.file_path
        if not db_path:
            return FetchResult(success=False, error="No database path provided")

        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.execute(config.query)
        rows = cursor.fetchall()
        records = [dict(row) for row in rows]
        conn.close()

        elapsed = (time.time() - start) * 1000
        return FetchResult(
            success=True,
            data=records,
            record_count=len(records),
            elapsed_ms=elapsed,
            metadata={"db": db_path, "query": config.query[:100]},
        )
    except Exception as exc:
        return FetchResult(
            success=False,
            error=str(exc),
            elapsed_ms=(time.time() - start) * 1000,
        )


def execute_data_source(config: DataSourceConfig) -> FetchResult:
    """Route to the correct connector based on source_type."""
    handlers = {
        "http": fetch_http,
        "rest": fetch_http,
        "api": fetch_http,
        "file": fetch_file,
        "csv": fetch_file,
        "database": fetch_database,
        "sqlite": fetch_database,
        "sql": fetch_database,
    }
    handler = handlers.get(config.source_type.lower())
    if handler is None:
        return FetchResult(
            success=False,
            error=f"Unknown source type: {config.source_type}. Supported: {', '.join(handlers)}",
        )
    return handler(config)


def format_as_table(records: list[dict[str, Any]], max_rows: int = 50) -> str:
    """Format records as a readable text table for notifications."""
    if not records:
        return "(empty)"

    headers = list(records[0].keys())
    rows = records[:max_rows]

    col_widths = {h: len(str(h)) for h in headers}
    for row in rows:
        for h in headers:
            col_widths[h] = max(col_widths[h], len(str(row.get(h, ""))))

    header_line = " | ".join(str(h).ljust(col_widths[h]) for h in headers)
    sep_line = "-+-".join("-" * col_widths[h] for h in headers)

    lines = [header_line, sep_line]
    for row in rows:
        line = " | ".join(str(row.get(h, "")).ljust(col_widths[h]) for h in headers)
        lines.append(line)

    if len(records) > max_rows:
        lines.append(f"... and {len(records) - max_rows} more rows")

    return "\n".join(lines)
